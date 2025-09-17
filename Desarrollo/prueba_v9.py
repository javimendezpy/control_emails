
''' Módulo para controlar la recepción de correos electrónicos de diferentes sistemas meteorológicos en Outlook y '
actualizar un archivo CSV con los resultados. '''

import pandas as pd 
import win32com.client
import re 
import os
import argparse
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

VERBOSE = True  # controla si se muestran los mensajes de tipo "info"

def log(msg, level="info", verbose=True):
    if not VERBOSE and verbose:
        return
    
    if level == "ok":
        print(f"[OK] {msg}")
    elif level == "error":
        print(f"[!!] {msg}")
    elif level == "info":
        print(f"[..] {msg}")
    elif level == "section":  # bloque grande para fechas
        print("\n" + "="*40)
        print(f"=== {msg} ===")
        print("="*40 + "\n")
    elif level == "subsection":  # bloque más pequeño para sistemas
        print(f"--- {msg} ---")


def conectar_outlook(nombre_cuenta: str, carpeta: str):

    ' Esta función conecta a Outlook y devuelve los mensajes de la carpeta especificada '
    
    log(f"Conectando a Outlook con la cuenta: {nombre_cuenta}, carpeta: {carpeta}", "info")
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        store = outlook.Stores[nombre_cuenta]
        bandeja_entrada = store.GetDefaultFolder(6) # Bandeja de entrada
        carpeta_dades_meteo = bandeja_entrada.Folders[carpeta] # Subcarpeta
        mensajes = carpeta_dades_meteo.Items
        mensajes.Sort("[ReceivedTime]", True)
        log("Conexión a Outlook establecida y carpeta encontrada.", "ok")
        return mensajes
    except Exception as e:
        log(f"No se pudo conectar a Outlook o acceder a la carpeta: {e}", "fail")
        raise


def extraer_remitente(remitente: str, id: str) -> str:

    ' Esta función extrae un identificador según el remitente y su id '

    if remitente == 'estaciones.meteo@dekra-industrial.es' and id == 'Olmillos_1':
        return 'estacionesmeteo (olmillos)'
    elif remitente == 'estaciones.meteo@dekra-industrial.es':
        return 'estacionesmeteo'
    elif remitente == 'windcubeinsights@vaisala.info':
        return 'windcube'
    elif remitente == 'emailrelay@konectgds.com':
        return 'emailrelay'
    elif remitente == 'status@support.zxlidars.com':
        return 'zx'
    return ''


def extraer_patron(remitente, id):

    ' Esta función extrae el patrón regex según el remitente y su id '

    ''' -------- Asuntos de ejemplo y formato de fechas para cada remitente --------
    estaciones.meteo -> LIDAR Punago-9_2025-08-12_00-10-00 ; YYYY-MM-DD_HH-MM-SS
    windcube -> WindCube Insights Fleet: New STA File from WLS71497  2025/07/31  00:10:00 ; YYYY/MM/DD  HH-MM-SS
    emailrelay -> LIDAR Villalube-6A_2025-08-11_00-10-00 ; YYYY-MM-DD_HH-MM-SS
    zx -> Daily Data: Wind10_1148@Y2025_M08_D02.CSV (Averaged data) ; YYYYY_MMM_DDD
    estaciones.meteo (Olmillos) -> Ammonit Data Logger Meteo-40M D243094 Olmillos_1  (signed) '''

    if remitente == 'estaciones.meteo@dekra-industrial.es' and id == 'Olmillos_1':
        return None # Olmillos no utiliza fecha en el asunto
    elif remitente == 'estaciones.meteo@dekra-industrial.es':
        return re.compile(rf"^{id}_(\d{{4}}-\d{{2}}-\d{{2}})_(\d{{2}}-\d{{2}}-\d{{2}})$")
    elif remitente == 'windcubeinsights@vaisala.info':
        return re.compile(rf"^WindCube Insights Fleet: New STA File from {id}\s+(\d{{4}}/\d{{2}}/\d{{2}})\s+(\d{{2}}[:\-]\d{{2}}[:\-]\d{{2}})$")
    elif remitente == 'emailrelay@konectgds.com':
        return re.compile(rf"^{id}_(\d{{4}}-\d{{2}}-\d{{2}})_(\d{{2}}-\d{{2}}-\d{{2}})$")
    elif remitente == 'status@support.zxlidars.com':
        return re.compile(rf"^Daily Data: Wind10_{id}@Y(\d{{4}})_M(\d{{2}})_D(\d{{2}})\.(?:CSV|ZPH) \(Averaged data\)$")
    return None


def extraer_fecha(asunto: str, patron, remitente: str, received_time=None) -> str | None:

    ' Esta función extrae la fecha del asunto del correo en función del identificador del remitente '

    ''' Devuelve la fecha de datos (YYYY-MM-DD) según remitente.
    Para estaciones.meteo/emailrelay → fecha del asunto -1 día
    Para windcube/zx → fecha del asunto directamente
    Para Olmillos → fecha de recepción -1 día
    '''

    if remitente == 'estacionesmeteo (olmillos)':
        if received_time is None:
            return None
        # Forzar datetime naive (sin tz)
        if hasattr(received_time, "replace"):
            received_time = received_time.replace(tzinfo=None)
        fecha = (pd.Timestamp(received_time) - pd.Timedelta(days=1)).date()
        return str(fecha)
    
    if not patron:
        return None
    
    m = patron.search(asunto)
    if not m:
        return None
    
    if remitente in ["estacionesmeteo", "emailrelay"]:
        fecha = pd.to_datetime(m.group(1)).date()-pd.Timedelta(days=1)
        return str(fecha)  # ya viene YYYY-MM-DD
    
    elif remitente == "windcube":
        fecha = pd.to_datetime(m.group(1).replace("/", "-")).date()
        return str(fecha)  # YYYY/MM/DD -> YYYY-MM-DD
    
    elif remitente == "zx":
        fecha = pd.to_datetime(f"{m.group(1)}-{m.group(2)}-{m.group(3)}").date()
        return str(fecha)
    return None


def filtrar_mensajes(mensajes, fecha_referencia: pd.Timestamp):

    ' Esta función filtra los mensajes de Outlook por fecha de recepción '

    dia_siguiente = fecha_referencia + pd.Timedelta(days=1)
    fecha_ini = pd.Timestamp(fecha_referencia).strftime('%d/%m/%Y 00:00 AM')
    fecha_fin = pd.Timestamp(dia_siguiente).strftime('%d/%m/%Y 23:59 PM')
    log(f"Filtrando mensajes por ReceivedTime entre: {fecha_ini} -> {fecha_fin}", "info")
    mensajes_filtrados = mensajes.Restrict(f"[ReceivedTime] >= '{fecha_ini}' AND [ReceivedTime] <= '{fecha_fin}'")
    try:
        count = mensajes_filtrados.Count
        log(f"Mensajes en rango: {count}", "ok")
    except Exception:
        log("No se pudo obtener la propiedad Count de 'mensajes_filtrados'.", "warn")
    return mensajes_filtrados


def procesar_sistemas(df, sistemas, mensajes_filtrados, fecha_referencia):

    ' Esta función procesa los sistemas y devuelve una lista de resultados marcando con 1 si se ha recibido correo y 0 si no '

    resultados = []

    for sistema in sistemas:

        remitente = df[df.iloc[:, 0] == sistema].iloc[0, 1]
        id = df[df.iloc[:, 0] == sistema].iloc[0, 2]

        log(f"Sistema: {sistema}", "subsection")

        rem = extraer_remitente(remitente, id)
        patron_asunto = extraer_patron(remitente, id)
        #print(patron_asunto)

        valor = 0  # Valor por defecto: no se recibió correo de ese remitente

        for msg in mensajes_filtrados:

            # print('msg.Subject: ', msg.Subject)
            # print('msg.ReceivedTime: ', msg.ReceivedTime)
            # print('msg.Sender: ', msg.Sender)

            try:
                sender = msg.Sender.GetExchangeUser().PrimarySmtpAddress
            except Exception:
                try:
                    sender = msg.SenderEmailAddress
                    # log(f"No se pudo obtener PrimarySmtpAddress; usando SenderEmailAddress: {sender}", "warn")
                except Exception:
                    # log("No se pudo obtener la dirección del remitente del mensaje.", "warn")
                    continue

            # sender = msg.SenderEmailAddress

            # print('Remitente: ', sender)

            if sender.lower() == remitente.lower():
                # print('Se ha encontrado un correo con remitente igual al del sistema: ', sender)
                fecha_asunto_tmp = extraer_fecha(
                    msg.Subject, 
                    patron_asunto, 
                    rem, 
                    received_time=msg.ReceivedTime
                )
                #print('Fecha asunto: ', fecha_asunto_tmp)
                if not fecha_asunto_tmp:
                    # log(f"Correo de {sender} encontrado pero el asunto no contiene la fecha esperada (o no coincide con patrón).", "warn")
                    continue

                fecha_asunto = pd.to_datetime(fecha_asunto_tmp).tz_localize(None).date()
                # log(f"Correo de {sender} -> fecha en asunto: {fecha_asunto}", "info")

                if fecha_asunto == fecha_referencia:
                    log(f"Correo recibido para {sistema}", "ok")
                    valor = 1
                    break
                else:
                    log(f"No se recibió correo para {sistema}", "error")

        print(f'\n--------------------------------------------------------------\n')        
        
        # Guardamos los resultados para cada sistema
        resultados.append({
            "Sistema": sistema,
            "Remitente": remitente, 
            "Fecha": fecha_referencia,
            "Valor": valor
        })

    return resultados


def actualizar_csv(resultados, fecha_referencia, output_file="control_emails.csv"):

    ' Esta función actualiza el archivo CSV con los resultados '
    
    fecha_col = str(fecha_referencia)

    log(f"Leyendo archivo '{output_file}'...", "info")
    tabla = pd.read_csv(output_file, index_col="Sistema")
    log(f"Archivo '{output_file}' leído. Sistemas en tabla: {len(tabla)}", "ok")

    if fecha_col not in tabla.columns:
        tabla[fecha_col] = 0
        log(f"Añadida columna {fecha_col} (inicializada a 0).", "info")

    for r in resultados:
        tabla.loc[r["Sistema"], fecha_col] = r["Valor"]

    cols_fijas = [c for c in tabla.columns if not re.match(r"^\d{4}-\d{2}-\d{2}$", c)]
    cols_fechas = sorted(
        [c for c in tabla.columns if re.match(r"^\d{4}-\d{2}-\d{2}$", c)],
        reverse=True
    )

    tabla = tabla[cols_fijas + cols_fechas]

    # Guardar
    tabla.to_csv(output_file)
    tabla.reset_index().to_excel("control_emails.xlsx", index=False)
    log(f"CSV '{output_file}' y Excel intermedio 'control_emails.xlsx' guardados.", "ok")

    # --- Aplicar formato condicional ---
    wb = load_workbook("control_emails.xlsx")
    ws = wb.active

    # Definir colores
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    rojo = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Detectar columnas de fechas (YYYY-MM-DD)
    columnas = [i+1 for i, cell in enumerate(ws[1]) if cell.value and re.match(r"\d{4}-\d{2}-\d{2}", str(cell.value))]
    log(f"Columnas de fecha detectadas para formato condicional: {len(columnas)}", "info")

    for col_idx in columnas:
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        # Regla 1 -> verde
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}",
                                      CellIsRule(operator='equal', formula=['1'], stopIfTrue=True, fill=verde))
        # Regla 0 -> rojo
        ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}",
                                      CellIsRule(operator='equal', formula=['0'], stopIfTrue=True, fill=rojo))

    wb.save("control_emails.xlsx")
    log("Formato condicional aplicado y 'control_emails.xlsx' guardado.", "ok")



def main(fecha_inicio: str, fecha_fin: str = None):

    ' Función principal que ejecuta el control de correos y actualización del CSV '
    log(f"Inicio del script. Fecha inicio: {fecha_inicio}, Fecha fin: {fecha_fin}", "info")

    log("Cargando 'control_emails.csv' para obtener la lista de sistemas...", "info")
    df = pd.read_csv('control_emails.csv')
    sistemas = df.iloc[:, 0].tolist()
    log(f"Sistemas cargados: {len(sistemas)}", "ok")

    mensajes = conectar_outlook(nombre_cuenta="energias.renovables.es@dekra.com", carpeta="Dades Meteo")
    log(f"Mensajes obtenidos desde Outlook (objeto).", "info")

    fecha_inicio_dt = pd.to_datetime(fecha_inicio).date()
    if fecha_fin:
        fecha_fin_dt = pd.to_datetime(fecha_fin).date()
    else:
        fecha_fin_dt = fecha_inicio_dt

    delta = (fecha_fin_dt - fecha_inicio_dt).days
    fechas_a_procesar = [fecha_inicio_dt + pd.Timedelta(days=i) for i in range(delta + 1)]
    log("Fechas a procesar: " + ", ".join([f.strftime("%Y-%m-%d") for f in fechas_a_procesar]), "info")


    for fecha_referencia in fechas_a_procesar:
        fecha_referencia = pd.to_datetime(fecha_referencia).date()
        log(f"Procesando fecha: {fecha_referencia}", "section")
        mensajes_filtrados = filtrar_mensajes(mensajes, fecha_referencia)
        resultados = procesar_sistemas(df, sistemas, mensajes_filtrados, fecha_referencia)
        actualizar_csv(resultados, fecha_referencia)
        log(f"Fecha {fecha_referencia} procesada", "ok")
        log("-"*40, "info")  # separador entre fechas
    
    log("Proceso completado para todas las fechas indicadas.", "done")

if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="Control de correos meteorológicos y actualización del CSV")
    parser.add_argument("fecha_inicio",type=str,help="Fecha inicial (YYYY-MM-DD)")
    parser.add_argument("fecha_fin",nargs="?",help="Fecha final (YYYY-MM-DD), opcional",default=None)
    args = parser.parse_args()

    log(f"Script lanzado con: fecha_inicio={args.fecha_inicio}, fecha_fin={args.fecha_fin}", "info")

    main(args.fecha_inicio, args.fecha_fin)